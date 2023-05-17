import openpyxl as openpyxl

wb=openpyxl.Workbook()

ws=wb.active

ws.title="Voce"

ws["A1"].value="ID"
ws["B1"].value="Naziv"
ws["C1"].value="Cena"

for i in range(2,7):
    ws.cell(row=i,column=1).value=i-1
    ws.cell(row=i,column=2).value=input("Unesite naziv voca ")
    ws.cell(row=i,column=3).value=eval(input("Unesite cenu voca "))

wb.save(filename="eksel3.xlsx") 