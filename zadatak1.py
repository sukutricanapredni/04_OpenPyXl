import openpyxl as openpyxl

wb=openpyxl.Workbook() ##Otvara excel workbook

ws=wb.active    ##Uzima aktivni sheet
ws.title='Naslov' ##menja naslov aktivnom sheetu

##Menjanje vrednosti odredjene celije-Nacin 1
ws['A1'].value="Ime"
ws['B1'].value="Prezime"
ws['A2'].value="Veljko"
ws['B2'].value="Krstic"

##Menjanje vrednosti odredjene celije-Nacin 2
ws.cell(row=3, column=1).value="Nikola"
ws.cell(row=3,column=2).value="Maric"



wb.save(filename='demo1.xlsx') ##Cuvamo workbook kao demo1.xlsx

