import openpyxl as openpyxl

wb=openpyxl.Workbook()

ws=wb.active

ws["A1"].value="Jedan"
c=ws["A1"]
c.font=openpyxl.styles.Font(name="Times New Roman",size=18,bold=True,italic=True)


wb.save(filename="eksel2.xlsx")