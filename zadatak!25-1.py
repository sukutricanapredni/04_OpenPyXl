import openpyxl as openpyxl

wb=openpyxl.load_workbook("eksel3.xlsx")

ws=wb.active

f=open("voce.txt","w")

for i in range(2,ws.max_row+1):
    print("ID:",ws.cell(row=i,column=1).value,file=f)
    print("Naziv:",ws.cell(row=i,column=2).value,file=f)
    print("Cena:",ws.cell(row=i,column=3).value,file=f)
    print("="*25,file=f)
    
f.close()
