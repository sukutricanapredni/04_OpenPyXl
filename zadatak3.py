import openpyxl as openpyxl

l=[line.strip() for line in open("proizvodi.txt")]
z=[[x.split("|")[0],int(x.split("|")[1]),eval(x.split("|")[2])] for x in l]

wb=openpyxl.Workbook()
ws=wb.active

ws['A1'].value="Naziv"
ws['B1'].value="Kolicina"
ws['C1'].value="Cena"

for i in range(2,len(z)+2):
    ws.cell(row=i,column=1).value=z[i-2][0]
    ws.cell(row=i,column=2).value=z[i-2][1]
    ws.cell(row=i,column=3).value=z[i-2][2]

wb.save(filename='proizvodi.xlsx')
