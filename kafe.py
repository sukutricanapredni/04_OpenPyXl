import psycopg2 as psycopg2
import openpyxl as openpyxl

print("1.Naziv-Cena\n2.ID-Naziv-Cena\n3.Naziv-Kofein\n4.Sve\n5.Naziv-Najskuplja\n6.Naziv-Najjaca")
print()
o=input("Izaberite opciju")


def query_sql(q):
        con=psycopg2.connect(
            database='kafe',
            user='postgres',
            port='5432',
            password='itoip',
            host='localhost'
        )

        cursor=con.cursor()

        cursor.execute(q)
        r=cursor.fetchall()
        cursor.close()
        con.close()
        return r


def naziv_cena_excel():
    l=query_sql('SELECT NAZIV,CENA FROM KAFE')
    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title="Naziv-Cena"

    ws["A1"].value="Naziv"
    ws["B1"].value="Cena"

    for i in range(2,len(l)+2):
          ws.cell(row=i,column=1).value=l[i-2][0]
          ws.cell(row=i,column=2).value=l[i-2][1]
    
    wb.save(filename="Naziv-Cena.xlsx")


def id_naziv_cena_excel():
    l=query_sql('SELECT ID_KAFE,NAZIV,CENA FROM KAFE')
    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title="ID-Naziv-Cena"

    ws["A1"].value="ID"
    ws["B1"].value="Naziv"
    ws["C1"].value="Cena"

    for i in range(2,len(l)+2):
          ws.cell(row=i,column=1).value=l[i-2][0]
          ws.cell(row=i,column=2).value=l[i-2][1]
          ws.cell(row=i,column=3).value=l[i-2][2]

    wb.save(filename="ID-Naziv-Cena.xlsx")

def naziv_kofein_excel():
    l=query_sql('SELECT NAZIV,KOFEIN FROM KAFE')
    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title="Naziv-Kofein"

    ws["A1"].value="Naziv"
    ws["B1"].value="Kofein"

    for i in range(2,len(l)+2):
          ws.cell(row=i,column=1).value=l[i-2][0]
          ws.cell(row=i,column=2).value=l[i-2][1]
    
    wb.save(filename="Naziv-Kofein.xlsx")

def id_naziv_cena_kofein_excel():
    l=query_sql('SELECT * FROM KAFE')
    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title="ID-Naziv-Cena-Kofein"

    ws["A1"].value="ID"
    ws["B1"].value="Naziv"
    ws["C1"].value="Cena"
    ws["D1"].value="Kofein"

    for i in range(2,len(l)+2):
          ws.cell(row=i,column=1).value=l[i-2][0]
          ws.cell(row=i,column=2).value=l[i-2][1]
          ws.cell(row=i,column=3).value=l[i-2][2]
          ws.cell(row=i,column=4).value=l[i-2][3]

    wb.save(filename="ID-Naziv-Cena-Kofein.xlsx")

def naziv_najskuplja():
    l=query_sql('SELECT NAZIV,CENA FROM KAFE WHERE CENA=(SELECT MAX(CENA) FROM KAFE)')
    print("Naziv:",l[0][0])
    print("Cena:",l[0][1])

def naziv_najjaca():
    l=query_sql('SELECT NAZIV,KOFEIN FROM KAFE WHERE KOFEIN=(SELECT MAX(KOFEIN) FROM KAFE)')
    print("Naziv:",l[0][0])
    print("Kofein:",l[0][1])

if o=="1":
     naziv_cena_excel()
elif o=="2":
     id_naziv_cena_excel()
elif o=="3":
     naziv_kofein_excel()
elif o=="4":
     id_naziv_cena_kofein_excel()
elif o=="5":
    naziv_najskuplja()
elif o=="6":
     naziv_najjaca()
else:
     print("Neispravan odabir")
