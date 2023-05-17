import openpyxl as openpyxl

wb=openpyxl.load_workbook("eksel1.xlsx")

ws=wb["Stranica 4"]

print(ws.max_column)
print(ws.max_row)

##1.Zadatak
##Odstampati sve podatke iz tabele od pocetka do kraja koristeci max row
m=ws.max_row

for i in range(2,m+1):
    print(ws.cell(row=i,column=1).value)
    print(ws.cell(row=i,column=2).value)