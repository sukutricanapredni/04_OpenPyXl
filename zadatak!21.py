import openpyxl as openpyxl
wb=openpyxl.Workbook()
ws=wb.active
ws.title="Stranica 1" #Naslov sheet-a

ws["A1"].value="Prva" ##Dodeljuje vrednost u celiju A1

wb.create_sheet("Stranica 2") ##Kreira novi sheet
ws=wb["Stranica 2"] ##Uzimamo drugi sheet po nazivu
ws["A1"].value="Druga" ##Upisujemo vrednost u celiju A1
wb.create_sheet("Stranica 3") ##Kreiramo novi sheet
ws=wb.worksheets[2] ##Uzimamo treci sheet po indeksu
ws["A1"].value="Treci" ##U celiju A1 upisujemo vrednost
print(wb.sheetnames)

wb.save(filename="eksel1.xlsx") ##cuvamo fajl