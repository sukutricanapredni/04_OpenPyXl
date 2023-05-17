import openpyxl as openpyxl

# ime_prezime=["Veljko Krstic","Nikola Maric","Sava Milunovic"]
# plata=[70000,80000,150000]


l=[line.strip() for line in open("zaposleni.txt")]
z=[[x.split("|")[0],eval(x.split("|")[1])] for x in l]

wb=openpyxl.Workbook()
ws=wb.active

ws['A1'].value="Ime Prezime"
ws.cell(row=1,column=2).value="Plata"

for i in range(2,len(z)+2):
    ws.cell(row=i,column=1).value=z[i-2][0]
    ws.cell(row=i,column=2).value=z[i-2][1]

wb.save(filename='zaposleni.xlsx')
