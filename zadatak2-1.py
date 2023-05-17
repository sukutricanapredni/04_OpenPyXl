import openpyxl as openpyxl

wb=openpyxl.load_workbook('zaposleni.xlsx')
ws=wb.active

# ##Citanje vrednosti iz odredjenih celija- Nacin 1
# ime_prezime=ws['A2'].value ##Uzima vrednost iz celije A2
# plata=ws['B2'].value ##Uzima vrednost iz celije B2

# ##Citanje vrednosti iz odredjenih celija- Nacin 2
# ime_prezime1=ws.cell(row=3,column=1).value
# plata1=ws.cell(row=3,column=2).value

# ##Citanje cele kolone A
# c=ws["A"]

# for i in c:
#     print(i.value)

# ##Citanje celog reda 2

# r=ws[2]

# for i in r:
#     print(i.value)

##Citanje vise celija odjednom
c=ws["A2":"B7"]

for i in c:
    print("Ime prezime:",i[0].value)
    print("Plata:",i[1].value)
    print("="*20)
    


# print("Ime prezime:",ime_prezime) 
# print("Plata:",plata)
# print("Ime prezime:",ime_prezime1) 
# print("Plata:",plata1)